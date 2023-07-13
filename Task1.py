from openpyxl import Workbook
from openpyxl import load_workbook
import pdb
from itertools import groupby

def generate_new_sheet(sheet_name, wb, wb_name, data):
    current_sn = wb.sheetnames

    if sheet_name not in current_sn:
        wb.create_sheet(sheet_name)
    else:
        del wb[sheet_name]
        # pdb.set_trace()
        wb.create_sheet(sheet_name)


    print(current_sn)
    current_sheet = wb[sheet_name]
    for e in data:
        current_sheet.append([e])

    # current_sheet.append(zy)

    return wb.save(filename=wb_name)


workbook = load_workbook(filename="name_of_plants_new.xlsx")

workbook.sheetnames
sheet = workbook["Worksheet"]

firstname_db = []
fullname_db = []
lastname_db = []

sheet["A"]
for val in sheet.iter_rows(min_row=2,values_only=True):
    x = val[0].split(" (")
    firstname_db.append(x[0])
    fullname_db.append(val[0])
    if ")" in x[-1] and x[-1] is not None:
        lastname_db.append("(" + x[-1])                    
    # lastname_db.append("(" + x[-1] if ")" in x[-1] else "")

print("fullname", fullname_db)
print("firstname", firstname_db)
print("lastname", lastname_db)


new_workbook = load_workbook(filename="list_combined.xlsx")
new_workbook.sheetnames
sheet = new_workbook["Sheet1"]

fullname_cd = []

sheet["A"]
for name in sheet.iter_rows(min_row=2,values_only=True):
    fullname_cd.append(name[0].replace(u'\xa0', u' ') if name[0] is not None else name[0])
 
print('full_name:', fullname_cd)

ffn_full = []

fullname_cd = list(set(fullname_cd))
fullname_db = list(set(fullname_db))

# for idx, x in enumerate(fullname_cd): 
#     print(f'No {idx}: {x}')
#     if x is not None:
#         for z in firstname_db:
#             if z is not None and z.lower() in x.lower():
#                 if z not in zy:
#                     zy.append(z)

# for idx, xz in enumerate(fullname_db):
#     # if idx == 0: print("Debugging xz");pdb.set_trace()
#     # if xz == "Abies concolor 'Hosta la Vista' (White Fir)": pdb.set_trace()
#     for idz,xy in enumerate(zy):
#         if xy.lower() in xz.lower():
#             FinalFullName.append(xz)

def full_name(used_name, combined_data, full_name_db):
    final_full_name=[]
    matched_name = []
    for idx, x in enumerate(combined_data):
        print(f'No {idx}: {x}')
        if x is not None:
            for z in used_name:
                if z is not None and z.lower() in x.lower():
                    if z not in matched_name:
                        matched_name.append(z)
    
    for idx, xz in enumerate(full_name_db):
    # if idx == 0: print("Debugging xz");pdb.set_trace()
    # if xz == "Abies concolor 'Hosta la Vista' (White Fir)": pdb.set_trace()
        for idz,xy in enumerate(matched_name):
            if xy.lower() in xz.lower():
                final_full_name.append(xz)
    return final_full_name
            
pdb.set_trace()

for idx,xf in enumerate(fullname_cd):
    if xf is not None:
    # if idx == 0: print("Debugging ab");pdb.set_trace()
        for idz,ab in enumerate(fullname_db):
            if ab is not None and ab.lower() in xf.lower():
                print("xf index:", idx)
                print("ab index:", idz)
                ffn_full.append(ab)

# fc={}
# dati= [fc.update({f'{k}': len(list(g))}) for k,g in groupby(p_id_2, lambda x: x)]
# papu=dict(filter(lambda k: k[1] > 1, fc.items()))

# print("last firstname:", firstname[0])
        
    # current_name = [x for x in full_name if x is not None and z.lower() in x.lower()]
    # matched_name_firs`t.append(current_name)
    # list(filter(lambda x: z.lower() in x.lower() if x is not None else x, full_name))
    
ffn_first = full_name(firstname_db, fullname_cd, fullname_db)
ffn_last = full_name(lastname_db, fullname_cd, fullname_db)

generate_new_sheet("With first name", workbook, "name_of_plants_new.xlsx", ffn_first)
generate_new_sheet("With last name", workbook, "name_of_plants_new.xlsx", ffn_last)
generate_new_sheet("With full name", workbook, "name_of_plants_new.xlsx", ffn_full)


# for z in lastname:
#     c_meto = list(filter(lambda y: z.lower() in y.lower() if y is not None else y, full_name))
#     print("Pass")
#     print(c_meto)
#     mnf.append(c_meto)

# c_meto = list(filter(lambda y: lastname[99].lower() in y.lower() if y is not None else y, full_name))
